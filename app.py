from datetime import date, datetime, timedelta, time
from pathlib import Path
from typing import Optional
import csv
import sqlite3
import tempfile
import hashlib
import json
import hmac
import io
import os
import re
import secrets
import shutil
import urllib.parse
import urllib.request

from fastapi import Depends, FastAPI, Form, HTTPException, Request, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from sqlalchemy import extract, func, or_, text
from sqlalchemy.orm import Session, joinedload

from database import Base, DB_PATH, IS_SQLITE, SessionLocal, engine, get_db
from models import (
    AppSetting,
    ActionHistory,
    Attestation,
    AttestationHistory,
    BlacklistEntry,
    Employee,
    Promotion,
    Interview,
    Punishment,
    SettingOption,
    User,
    Vacation,
)

BASE_DIR = Path(__file__).resolve().parent
BACKUP_DIR = BASE_DIR / "backups"
EXPORT_DIR = BASE_DIR / "exports"
BACKUP_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

app = FastAPI(title="HolyFake Staff Control Panel")
AUTH_SECRET = os.getenv("HF_SECRET_KEY", "holyfake-local-secret-change-me")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

ROLE_LABELS = {
    "owner": "Главный администратор",
    "admin": "Администратор",
    "moderator": "Руководитель",
    "staff": "Сотрудник",
}
ROLE_ORDER = {"staff": 1, "moderator": 2, "admin": 3, "owner": 4}
WRITE_ROLES = {"owner", "admin", "moderator"}
ADMIN_ROLES = {"owner", "admin"}
OWNER_ROLES = {"owner"}

SECTION_PERMISSIONS = [
    {"key": "view_dashboard", "page": "dashboard", "url": "/", "icon": "⌂", "label": "Главная"},
    {"key": "view_employees", "page": "employees", "url": "/employees", "icon": "👥", "label": "Сотрудники"},
    {"key": "view_punishments", "page": "punishments", "url": "/punishments", "icon": "⚠", "label": "Выдача наказаний"},
    {"key": "view_removals", "page": "removals", "url": "/removals", "icon": "⏱", "label": "Сроки снятия"},
    {"key": "view_vacations", "page": "vacations", "url": "/vacations", "icon": "🌴", "label": "Отпуска"},
    {"key": "view_promotions", "page": "promotions", "url": "/promotions", "icon": "⬆", "label": "Повышения"},
    {"key": "view_attestation", "page": "attestation", "url": "/attestation", "icon": "🎓", "label": "Аттестация"},
    {"key": "view_interviews", "page": "interviews", "url": "/interviews", "icon": "📝", "label": "Собеседование"},
    {"key": "view_chskp", "page": "chskp", "url": "/chskp", "icon": "⛔", "label": "ЧСКП"},
    {"key": "view_chsp", "page": "chsp", "url": "/chsp", "icon": "🚫", "label": "ЧСП"},
    {"key": "view_archive", "page": "archive", "url": "/archive", "icon": "🗄", "label": "Архив"},
    {"key": "view_notifications", "page": "notifications", "url": "/notifications", "icon": "🔔", "label": "Уведомления"},
    {"key": "view_history", "page": "history", "url": "/history", "icon": "📜", "label": "Журнал действий"},
    {"key": "view_import_export", "page": "import_export", "url": "/import-export", "icon": "⇅", "label": "Импорт / экспорт"},
    {"key": "view_backups", "page": "backups", "url": "/backups", "icon": "💾", "label": "Резервные копии"},
    {"key": "view_settings", "page": "settings", "url": "/settings", "icon": "⚙", "label": "Настройки"},
]
ACTION_PERMISSIONS = [
    ("manage_employees", "Сотрудники: добавлять, редактировать, снимать и возвращать"),
    ("manage_punishments", "Наказания: выдавать и снимать вручную"),
    ("manage_vacations", "Отпуска: выдавать и завершать"),
    ("manage_promotions", "Повышения: повышать сотрудников"),
    ("manage_attestation", "Аттестация: назначать, принимать и отклонять"),
    ("manage_interviews", "Собеседование: добавлять и удалять анкеты"),
    ("manage_chskp", "ЧСКП: вносить и удалять записи"),
    ("manage_chsp", "ЧСП: вносить и удалять записи"),
    ("manage_import_export", "Импорт / экспорт: загружать и скачивать таблицы"),
    ("manage_backups", "Резервные копии: создавать и восстанавливать базу"),
    ("manage_settings", "Настройки: менять списки и сроки"),
    ("manage_users", "Доступ к сайту: создавать, блокировать, менять роли и права"),
]
ALL_PERMISSION_KEYS = [item["key"] for item in SECTION_PERMISSIONS] + [key for key, _ in ACTION_PERMISSIONS]
PERMISSION_GROUPS = [
    ("Какие разделы видит", [(item["key"], item["label"]) for item in SECTION_PERMISSIONS]),
    ("Что может делать", ACTION_PERMISSIONS),
]
ROLE_DEFAULT_PERMISSIONS = {
    "owner": set(ALL_PERMISSION_KEYS),
    "admin": set(ALL_PERMISSION_KEYS) - {"manage_users"},
    "moderator": {
        "view_dashboard", "view_employees", "view_punishments", "view_removals", "view_vacations", "view_promotions",
        "view_attestation", "view_interviews", "view_chskp", "view_chsp", "view_archive", "view_notifications", "view_history",
        "manage_employees", "manage_punishments", "manage_vacations", "manage_promotions", "manage_attestation", "manage_interviews", "manage_chskp", "manage_chsp",
    },
    "staff": {"view_dashboard", "view_employees", "view_notifications"},
}
SESSION_TTL_SECONDS = 60 * 60 * 12
LOGIN_LIMIT = 5
LOGIN_WINDOW_SECONDS = 60 * 10
LOGIN_LOCK_SECONDS = 60 * 10
LOGIN_ATTEMPTS: dict[str, dict] = {}

DEFAULT_OPTIONS = {
    # Чистая база: тестовых ников нет. Должности и статусы остаются.
    "positions": ["Владелец", "Гл.Админ", "Админ", "Ст.сотрудник", "Спектатор", "Вед.сотрудник", "Сотрудник", "Мл.сотрудник", "Стажер"],
    "accepted_by": [],
    "punishment_issuers": ["Система"],
    "interviewers": [],
    "statuses": ["Активен", "В отпуске", "Снят"],
}

POSITION_MIGRATION = {
    "Администратор": "Админ",
    "Старший Сотрудник": "Ст.сотрудник",
    "Мл. Сотрудник": "Мл.сотрудник",
    "Старший Модератор": "Ст.сотрудник",
    "Модератор": "Вед.сотрудник",
}

MANUAL_LEVELS = ["Мл.сотрудник", "Сотрудник", "Вед.сотрудник", "Спектатор"]
POSITION_RANK = {
    "Стажер": 1,
    "Мл.сотрудник": 2,
    "Сотрудник": 3,
    "Вед.сотрудник": 4,
    "Спектатор": 5,
    "Ст.сотрудник": 6,
    "Админ": 7,
    "Гл.Админ": 8,
    "Владелец": 9,
}
POSITION_ORDER = {
    "Владелец": 0,
    "Гл.Админ": 1,
    "Админ": 2,
    "Ст.сотрудник": 3,
    "Спектатор": 4,
    "Вед.сотрудник": 5,
    "Сотрудник": 6,
    "Мл.сотрудник": 7,
    "Стажер": 8,
    "-": 99,
    "": 99,
}
POSITION_CSS = {
    "Владелец": "pos-owner",
    "Гл.Админ": "pos-chief-admin",
    "Админ": "pos-admin",
    "Ст.сотрудник": "pos-senior",
    "Спектатор": "pos-spectator",
    "Вед.сотрудник": "pos-lead",
    "Сотрудник": "pos-staff",
    "Мл.сотрудник": "pos-junior",
    "Стажер": "pos-trainee",
    "-": "pos-empty",
    "": "pos-empty",
}


def position_sort_number(position: str) -> int:
    return POSITION_ORDER.get(position or "", 99)


def position_sort_key(employee: Employee):
    return (position_sort_number(employee.position), (employee.nick or "").lower())


def position_css_class(position: str) -> str:
    return POSITION_CSS.get(position or "", "pos-empty")


def position_label(position: str) -> str:
    labels = {
        "Владелец": "👑 Владелец",
        "Гл.Админ": "👑 Гл.Админ",
        "Админ": "🛡 Админ",
    }
    return labels.get(position or "", position or "-")

ATTESTATION_TARGETS = {
    "Стажер": ["Мл.сотрудник"],
    "Мл.сотрудник": ["Сотрудник"],
    "Сотрудник": ["Вед.сотрудник"],
    "Вед.сотрудник": ["Ст.сотрудник", "Спектатор"],
    "Ст.сотрудник": [],
    "Спектатор": [],
    "Админ": [],
    "Гл.Админ": [],
    "Владелец": [],
}
IMPORTANT_CONFIRM_TEXT = "Вы уверены? Действие будет сохранено в журнале."
BLACKLIST_LABELS = {"chskp": "ЧСКП", "chsp": "ЧСП"}
BLACKLIST_FULL = {"ЧСКП": "черный список команды проекта", "ЧСП": "черный список проекта"}


# -------------------------- common helpers --------------------------

def parse_date(value: Optional[str], fallback: Optional[date] = None) -> Optional[date]:
    if not value:
        return fallback
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_date_flexible(value: str, fallback: Optional[date] = None) -> Optional[date]:
    value = (value or "").strip()
    if not value:
        return fallback
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return fallback


def parse_bool_value(value: Optional[str], default: bool = False) -> bool:
    text = norm(value)
    if text in {"1", "true", "yes", "да", "есть", "активен", "активно", "✓", "✅", "галочка", "выдан", "выдано", "доступ есть"}:
        return True
    if text in {"0", "false", "no", "нет", "-", "x", "х", "✗", "❌", "крестик", "не выдан", "не выдано", "доступа нет"}:
        return False
    return default


def yesno(value: bool) -> str:
    return "✅" if value else "❌"


def manual_due(employee: Employee) -> bool:
    if employee.is_archived or employee.manual_access or employee.position != "Стажер" or not employee.accepted_date:
        return False
    return employee.accepted_date <= date.today() - timedelta(days=2)


def parse_time(value: Optional[str]) -> time:
    if not value:
        return datetime.now().time().replace(microsecond=0, second=0)
    return datetime.strptime(value, "%H:%M").time()


def redirect_to(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=303)


def render(request: Request, name: str, context: dict):
    current_user = getattr(request.state, "current_user", None)
    can_perm = lambda key: has_permission(current_user, key)
    context.setdefault("request", request)
    context.setdefault("current_user", current_user)
    context.setdefault("role_labels", ROLE_LABELS)
    context.setdefault("can_perm", can_perm)
    context.setdefault("nav_items", visible_nav_items(current_user))
    context.setdefault("can_write", bool(current_user and any(k.startswith("manage_") for k in user_permission_set(current_user))))
    context.setdefault("can_admin", bool(current_user and (current_user.role in ADMIN_ROLES or has_permission(current_user, "view_settings"))))
    context.setdefault("can_owner", bool(current_user and current_user.role in OWNER_ROLES))
    return templates.TemplateResponse(name, context)


def get_setting(db: Session, key: str, default: str) -> str:
    item = db.get(AppSetting, key)
    return item.value if item else default


def get_int_setting(db: Session, key: str, default: int) -> int:
    try:
        return int(get_setting(db, key, str(default)))
    except ValueError:
        return default


def set_setting(db: Session, key: str, value: str) -> None:
    item = db.get(AppSetting, key)
    if item:
        item.value = value
    else:
        db.add(AppSetting(key=key, value=value))


def option_values(db: Session, category: str) -> list[str]:
    rows = (
        db.query(SettingOption)
        .filter(SettingOption.category == category)
        .order_by(SettingOption.sort_order.asc(), SettingOption.value.asc())
        .all()
    )
    return [row.value for row in rows]


def all_options(db: Session) -> dict[str, list[str]]:
    return {
        "positions": option_values(db, "positions"),
        "accepted_by": option_values(db, "accepted_by"),
        "punishment_issuers": option_values(db, "punishment_issuers"),
        "interviewers": option_values(db, "interviewers"),
        "statuses": option_values(db, "statuses"),
    }


def add_history(db: Session, employee_id: Optional[int], action_type: str, details: str, actor: str = "Система") -> None:
    db.add(ActionHistory(employee_id=employee_id, action_type=action_type, details=details, actor=actor or "Система"))


def eligible_interviewers(db: Session) -> list[Employee]:
    return (
        db.query(Employee)
        .filter(Employee.is_archived.is_(False))
        .order_by(Employee.position.asc(), Employee.nick.asc())
        .all()
    )


def is_senior_enough(employee: Employee) -> bool:
    return POSITION_RANK.get(employee.position, 0) >= POSITION_RANK["Ст.сотрудник"]


def allowed_attestation_targets(position: str) -> list[str]:
    return ATTESTATION_TARGETS.get(position, [])


def validate_attestation_target(employee: Employee, target_position: str) -> None:
    allowed = allowed_attestation_targets(employee.position)
    if target_position not in allowed:
        if POSITION_RANK.get(target_position, 0) <= POSITION_RANK.get(employee.position, 0):
            raise HTTPException(400, "Нельзя назначить аттестацию на должность ниже или равную текущей")
        raise HTTPException(400, "Для текущей должности доступна только следующая ступень аттестации по иерархии")


def remaining_days_raw(target: Optional[date]) -> int:
    if not target:
        return 0
    return (target - date.today()).days


def remaining_days(target: Optional[date]) -> int:
    return max(remaining_days_raw(target), 0)


def status_badge_class(status: str) -> str:
    lowered = (status or "").lower()
    if "пройден" in lowered:
        return "badge-green"
    if "снято" in lowered or "завершен" in lowered or "истек" in lowered:
        return "badge-muted"
    if "отпуск" in lowered:
        return "badge-blue"
    if "актив" in lowered:
        return "badge-green"
    if "кд" in lowered or "задерж" in lowered or "провер" in lowered:
        return "badge-yellow"
    if "чскп" in lowered or "снят" in lowered:
        return "badge-red"
    return "badge-orange"


def due_class(target: Optional[date], status: str = "") -> str:
    lowered = (status or "").lower()
    if "снято" in lowered or "завершен" in lowered or "истек" in lowered:
        return "due-done"
    if not target:
        return ""
    days = (target - date.today()).days
    if days < 0:
        return "due-expired"
    if days <= 2:
        return "due-soon"
    return "due-active"


def employee_counts(db: Session) -> dict[int, dict[str, int]]:
    data: dict[int, dict[str, int]] = {}
    active = db.query(Punishment).filter(Punishment.status == "Активно").all()
    for p in active:
        item = data.setdefault(p.employee_id, {"warnings": 0, "reprimands": 0})
        if p.type == "Предупреждение":
            item["warnings"] += 1
        elif p.type == "Выговор":
            item["reprimands"] += 1
    return data


def norm(value: Optional[str]) -> str:
    return (value or "").strip().lower()


def blacklist_status(entry: BlacklistEntry) -> str:
    if entry.term_type == "Бессрочно" or not entry.expires_at:
        return "Активно"
    return "Активно" if entry.expires_at >= date.today() else "Срок истек"


def chskp_status(entry: BlacklistEntry) -> str:
    # Оставлено для старых шаблонов, теперь используется общий blacklist_status.
    return blacklist_status(entry)


def active_blacklist_query(db: Session, list_type: Optional[str] = None):
    today = date.today()
    query = db.query(BlacklistEntry).filter(
        or_(BlacklistEntry.term_type == "Бессрочно", BlacklistEntry.expires_at.is_(None), BlacklistEntry.expires_at >= today)
    )
    if list_type:
        query = query.filter(BlacklistEntry.list_type == list_type)
    return query


def active_chskp_query(db: Session):
    return active_blacklist_query(db, "ЧСКП")


def active_chsp_query(db: Session):
    return active_blacklist_query(db, "ЧСП")


def find_blacklist_matches(
    db: Session,
    nick: str = "",
    discord: str = "",
    discord_id: str = "",
    telegram: str = "",
    telegram_id: str = "",
    email: str = "",
    list_type: Optional[str] = None,
) -> list[dict]:
    incoming = {
        "nick": ("Ник", norm(nick)),
        "discord": ("Дискорд", norm(discord)),
        "discord_id": ("Discord ID", norm(discord_id)),
        "telegram": ("Telegram", norm(telegram)),
        "telegram_id": ("Telegram ID", norm(telegram_id)),
        "email": ("Почта", norm(email)),
    }
    matches: list[dict] = []
    for entry in active_blacklist_query(db, list_type).all():
        fields = []
        for attr, (label, value) in incoming.items():
            if value and value == norm(getattr(entry, attr, "")):
                fields.append(label)
        if fields:
            matches.append({"entry": entry, "fields": fields})
    return matches


def blacklist_action_label(matches: list[dict]) -> str:
    labels = sorted({getattr(m["entry"], "list_type", "ЧСКП") for m in matches})
    return "/".join(labels) if labels else "ЧС"


def notifications_data(db: Session) -> dict[str, list]:
    maintenance_update(db)
    today = date.today()
    soon_limit = today + timedelta(days=3)
    punishment_items = (
        db.query(Punishment)
        .join(Employee)
        .filter(Punishment.status == "Активно", Punishment.remove_date <= soon_limit, Employee.is_archived.is_(False))
        .order_by(Punishment.remove_date.asc())
        .all()
    )
    vacation_items = (
        db.query(Vacation)
        .join(Employee)
        .filter(Vacation.status == "Активно", Vacation.end_date <= soon_limit, Employee.is_archived.is_(False))
        .order_by(Vacation.end_date.asc())
        .all()
    )
    return {"punishments": punishment_items, "vacations": vacation_items}


# Jinja helpers
templates.env.globals["remaining_days"] = remaining_days
templates.env.globals["remaining_days_raw"] = remaining_days_raw
templates.env.globals["badge_class"] = status_badge_class
templates.env.globals["due_class"] = due_class
templates.env.globals["today"] = date.today
templates.env.globals["chskp_status"] = chskp_status
templates.env.globals["blacklist_status"] = blacklist_status
templates.env.globals["position_css_class"] = position_css_class
templates.env.globals["position_sort_number"] = position_sort_number
templates.env.globals["position_label"] = position_label
templates.env.globals["yesno"] = yesno
templates.env.globals["manual_due"] = manual_due
templates.env.globals["manual_levels"] = MANUAL_LEVELS


# -------------------------- auth helpers --------------------------

def hash_password(password: str, salt: Optional[str] = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 180_000)
    return f"pbkdf2_sha256${salt}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, salt, digest = stored_hash.split("$", 2)
        if algorithm != "pbkdf2_sha256":
            return False
        candidate = hash_password(password, salt).split("$", 2)[2]
        return hmac.compare_digest(candidate, digest)
    except Exception:
        return False


def make_auth_token(user_id: int) -> str:
    issued_at = str(int(datetime.now().timestamp()))
    nonce = secrets.token_urlsafe(12)
    value = f"{user_id}:{issued_at}:{nonce}"
    signature = hmac.new(AUTH_SECRET.encode("utf-8"), value.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{value}:{signature}"


def parse_auth_token(token: str) -> Optional[int]:
    try:
        user_id, issued_at, nonce, signature = token.split(":", 3)
        value = f"{user_id}:{issued_at}:{nonce}"
        expected = hmac.new(AUTH_SECRET.encode("utf-8"), value.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None
        if int(datetime.now().timestamp()) - int(issued_at) > SESSION_TTL_SECONDS:
            return None
        return int(user_id)
    except Exception:
        return None


def user_public(user: User) -> dict:
    return {
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "role": user.role,
        "role_label": ROLE_LABELS.get(user.role, user.role),
        "status": user.status,
        "permissions": sorted(user_permission_set(user)),
        "last_login_at": user.last_login_at.isoformat() if user.last_login_at else None,
        "created_at": user.created_at.isoformat() if user.created_at else None,
    }


def wants_json(request: Request) -> bool:
    accept = request.headers.get("accept", "")
    api_auth_paths = ("/auth/me",)
    return request.url.path.startswith("/api") or request.url.path in api_auth_paths or "application/json" in accept


def login_error_redirect(code: str) -> RedirectResponse:
    return redirect_to(f"/login?error={urllib.parse.quote(code)}")


def set_auth_cookie(response: RedirectResponse | JSONResponse, user_id: int) -> None:
    response.set_cookie(
        "hf_auth",
        make_auth_token(user_id),
        httponly=True,
        samesite="lax",
        max_age=SESSION_TTL_SECONDS,
    )


def clear_auth_cookie(response: RedirectResponse | JSONResponse) -> None:
    response.delete_cookie("hf_auth")


def role_at_least(user: User, minimum: str) -> bool:
    return ROLE_ORDER.get(user.role, 0) >= ROLE_ORDER.get(minimum, 999)


def load_custom_permissions(user: User) -> Optional[set[str]]:
    raw = (getattr(user, "permissions", "") or "").strip()
    if not raw:
        return None
    try:
        data = json.loads(raw)
        if isinstance(data, list):
            return {str(x) for x in data if str(x) in ALL_PERMISSION_KEYS}
    except Exception:
        return None
    return None


def user_permission_set(user: Optional[User]) -> set[str]:
    if not user:
        return set()
    if user.status != "active":
        return set()
    if user.role == "owner":
        return set(ALL_PERMISSION_KEYS)
    custom = load_custom_permissions(user)
    if custom is not None:
        return custom
    return set(ROLE_DEFAULT_PERMISSIONS.get(user.role, set()))


def has_permission(user: Optional[User], permission: str) -> bool:
    if not user:
        return False
    if user.role == "owner" and user.status == "active":
        return True
    return permission in user_permission_set(user)


def require_permission(request: Request, permission: str) -> User:
    user = require_user(request)
    if not has_permission(user, permission):
        raise HTTPException(status_code=403, detail="Нет доступа к этому разделу или действию")
    return user


def visible_nav_items(user: Optional[User]) -> list[dict]:
    return [item for item in SECTION_PERMISSIONS if has_permission(user, item["key"])]


def editable_user_permissions(user: User) -> list[str]:
    return sorted(user_permission_set(user))


def save_user_permissions(user: User, permissions: list[str]) -> None:
    clean = sorted({p for p in permissions if p in ALL_PERMISSION_KEYS})
    user.permissions = json.dumps(clean, ensure_ascii=False)


def require_user(request: Request) -> User:
    user = getattr(request.state, "current_user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Нужно войти в аккаунт")
    if getattr(user, "status", "active") != "active":
        raise HTTPException(status_code=401, detail="Аккаунт заблокирован")
    return user


def require_write(request: Request) -> User:
    user = require_user(request)
    if user.role not in WRITE_ROLES:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    return user


def require_admin(request: Request) -> User:
    user = require_user(request)
    if user.role not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Нет доступа к административному разделу")
    return user


def require_owner(request: Request) -> User:
    user = require_user(request)
    if user.role not in OWNER_ROLES:
        raise HTTPException(status_code=403, detail="Только главный администратор может выполнить это действие")
    return user


def login_key(request: Request, identifier: str) -> str:
    ip = request.client.host if request.client else "local"
    return f"{ip}:{identifier.lower().strip()}"


def login_locked_until(request: Request, identifier: str) -> Optional[datetime]:
    item = LOGIN_ATTEMPTS.get(login_key(request, identifier))
    if not item:
        return None
    locked_until = item.get("locked_until")
    if locked_until and locked_until > datetime.now():
        return locked_until
    if locked_until and locked_until <= datetime.now():
        LOGIN_ATTEMPTS.pop(login_key(request, identifier), None)
    return None


def register_login_failure(request: Request, identifier: str) -> None:
    key = login_key(request, identifier)
    now = datetime.now()
    item = LOGIN_ATTEMPTS.setdefault(key, {"attempts": [], "locked_until": None})
    item["attempts"] = [t for t in item.get("attempts", []) if (now - t).total_seconds() <= LOGIN_WINDOW_SECONDS]
    item["attempts"].append(now)
    if len(item["attempts"]) >= LOGIN_LIMIT:
        item["locked_until"] = now + timedelta(seconds=LOGIN_LOCK_SECONDS)


def clear_login_failures(request: Request, identifier: str) -> None:
    LOGIN_ATTEMPTS.pop(login_key(request, identifier), None)


def json_or_redirect(request: Request, payload: dict, redirect_url: str, status_code: int = 200, set_cookie_user_id: Optional[int] = None):
    if wants_json(request):
        response = JSONResponse(payload, status_code=status_code)
    else:
        response = redirect_to(redirect_url)
    if set_cookie_user_id:
        set_auth_cookie(response, set_cookie_user_id)
    return response


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if wants_json(request):
        return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)
    if exc.status_code == 401:
        response = redirect_to("/login?error=session")
        clear_auth_cookie(response)
        return response
    denied_user = getattr(request.state, "current_user", None)
    denied_context = {
        "request": request,
        "detail": exc.detail,
        "current_user": denied_user,
        "role_labels": ROLE_LABELS,
        "can_perm": lambda key: has_permission(denied_user, key),
        "nav_items": visible_nav_items(denied_user),
    }
    if exc.status_code == 403:
        return templates.TemplateResponse("access_denied.html", denied_context, status_code=403)
    return templates.TemplateResponse("access_denied.html", denied_context, status_code=exc.status_code)


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    public_prefixes = ("/static", "/login", "/auth/login", "/health")
    path = request.url.path
    if path.startswith(public_prefixes):
        return await call_next(request)

    user_id = parse_auth_token(request.cookies.get("hf_auth", ""))
    if not user_id:
        if wants_json(request):
            return JSONResponse({"detail": "Нужно войти в аккаунт"}, status_code=401)
        return redirect_to("/login")

    with SessionLocal() as db:
        user = db.get(User, user_id)
        if not user:
            response = JSONResponse({"detail": "Сессия недействительна"}, status_code=401) if wants_json(request) else redirect_to("/login?error=session")
            clear_auth_cookie(response)
            return response
        if getattr(user, "status", "active") != "active":
            response = JSONResponse({"detail": "Аккаунт заблокирован"}, status_code=403) if wants_json(request) else login_error_redirect("blocked")
            clear_auth_cookie(response)
            return response
        request.state.current_user = user
        response = await call_next(request)
        return response


@app.middleware("http")
async def auto_maintenance(request: Request, call_next):
    if request.method == "GET" and not request.url.path.startswith(("/static", "/login", "/auth/login", "/health")):
        with SessionLocal() as db:
            maintenance_update(db)
    return await call_next(request)


# -------------------------- maintenance / seed --------------------------

def maintenance_update(db: Session) -> None:
    """Автоматически снимает просроченные наказания и завершает отпуска."""
    today_value = date.today()
    changed = False

    active_punishments = db.query(Punishment).filter(Punishment.status == "Активно").all()
    for p in active_punishments:
        if p.remove_date <= today_value:
            p.status = "Снято автоматически"
            p.removed_at = datetime.now()
            add_history(
                db,
                p.employee_id,
                "Автоснятие наказания",
                f"{p.type} снято автоматически. Причина выдачи: {p.reason}",
                "Система",
            )
            changed = True

    active_vacations = db.query(Vacation).filter(Vacation.status == "Активно").all()
    for v in active_vacations:
        if v.end_date <= today_value:
            v.status = "Завершен автоматически"
            employee = db.get(Employee, v.employee_id)
            if employee and not employee.is_archived and employee.status == "В отпуске":
                has_other_active = (
                    db.query(Vacation)
                    .filter(Vacation.employee_id == employee.id, Vacation.id != v.id, Vacation.status == "Активно")
                    .count()
                )
                if not has_other_active:
                    employee.status = "Активен"
            add_history(db, v.employee_id, "Автозавершение отпуска", f"Отпуск завершен автоматически: {v.reason}", "Система")
            changed = True

    cooldown_attestations = db.query(Attestation).filter(Attestation.status == "КД", Attestation.cooldown_until <= today_value).all()
    for a in cooldown_attestations:
        a.status = "Активно"
        a.cooldown_until = None
        add_history(db, a.employee_id, "Аттестация", "КД 2 дня закончился, можно сдавать повторно", "Система")
        changed = True

    delayed_attestations = db.query(Attestation).filter(Attestation.status == "Задержка", Attestation.delay_until <= today_value).all()
    for a in delayed_attestations:
        a.status = "Активно"
        a.remaining_attempts = 3
        a.delay_until = None
        a.cooldown_until = None
        add_history(db, a.employee_id, "Аттестация", "Задержка закончилась, попытки восстановлены до 3", "Система")
        changed = True

    if changed:
        db.commit()


def reset_options(db: Session, category: str, values: list[str]) -> None:
    db.query(SettingOption).filter(SettingOption.category == category).delete()
    for i, value in enumerate(values):
        db.add(SettingOption(category=category, value=value, sort_order=i))


def seed_data(db: Session) -> None:
    # Один раз обновляем дефолтные списки под последние требования пользователя.
    if get_setting(db, "defaults_v3_applied", "0") != "1":
        for category, values in DEFAULT_OPTIONS.items():
            reset_options(db, category, values)
        for old, new in POSITION_MIGRATION.items():
            db.query(Employee).filter(Employee.position == old).update({Employee.position: new})
        db.query(Employee).filter(Employee.status == "На проверке").update({Employee.status: "Активен"})
        set_setting(db, "defaults_v3_applied", "1")
        db.commit()

    # Обновляем только список должностей под новую иерархию, не трогая остальные настройки.
    if get_setting(db, "positions_hierarchy_v4_applied", "0") != "1":
        reset_options(db, "positions", DEFAULT_OPTIONS["positions"])
        set_setting(db, "positions_hierarchy_v4_applied", "1")
        db.commit()

    # Если база совсем чистая.
    for category, values in DEFAULT_OPTIONS.items():
        if not db.query(SettingOption).filter(SettingOption.category == category).first():
            for i, value in enumerate(values):
                db.add(SettingOption(category=category, value=value, sort_order=i))

    if not db.get(AppSetting, "warning_days"):
        set_setting(db, "warning_days", "7")
    if not db.get(AppSetting, "reprimand_days"):
        set_setting(db, "reprimand_days", "14")

    # Первый системный доступ. Создается только если его еще нет.
    # Важно: существующие пароли/роли/блокировки не перезаписываются при перезапуске.
    owner = db.query(User).filter(func.lower(User.username) == "owner").first()
    if not owner:
        db.add(User(username="owner", email="owner@holyfake.local", password_hash=hash_password("HolyFake#2026!"), role="owner", status="active"))

    # Мягкая миграция старой роли viewer -> staff, без сброса пароля.
    db.query(User).filter(User.role == "viewer").update({User.role: "staff"})


    # Финальная чистая сборка: тестовые сотрудники, ЧСКП/ЧСП, собеседования и т.д. не создаются.
    # В базе остаются только системный owner и редактируемые списки должностей/статусов.

    db.commit()


def ensure_schema() -> None:
    """Мини-миграции для локальной SQLite: добавляет новые колонки в старые базы."""
    if not IS_SQLITE:
        return

    with engine.begin() as conn:
        user_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(users)").fetchall()}
        if user_cols:
            if "email" not in user_cols:
                conn.exec_driver_sql("ALTER TABLE users ADD COLUMN email VARCHAR(160) DEFAULT ''")
                conn.exec_driver_sql("UPDATE users SET email = username || '@holyfake.local' WHERE email IS NULL OR email = ''")
            if "status" not in user_cols:
                conn.exec_driver_sql("ALTER TABLE users ADD COLUMN status VARCHAR(30) DEFAULT 'active'")
            if "permissions" not in user_cols:
                conn.exec_driver_sql("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT ''")
            if "last_login_at" not in user_cols:
                conn.exec_driver_sql("ALTER TABLE users ADD COLUMN last_login_at DATETIME")
            if "updated_at" not in user_cols:
                conn.exec_driver_sql("ALTER TABLE users ADD COLUMN updated_at DATETIME")
            conn.exec_driver_sql("UPDATE users SET role = 'staff' WHERE role = 'viewer'")
            conn.exec_driver_sql("UPDATE users SET status = 'active' WHERE status IS NULL OR status = ''")

        blacklist_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(blacklist_entries)").fetchall()}
        if blacklist_cols and "list_type" not in blacklist_cols:
            conn.exec_driver_sql("ALTER TABLE blacklist_entries ADD COLUMN list_type VARCHAR(20) DEFAULT 'ЧСКП'")
            conn.exec_driver_sql("UPDATE blacklist_entries SET list_type = 'ЧСКП' WHERE list_type IS NULL OR list_type = ''")

        employee_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(employees)").fetchall()}
        if "two_fa_enabled" not in employee_cols:
            conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN two_fa_enabled BOOLEAN DEFAULT 0")
        if "manual_access" not in employee_cols:
            conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN manual_access BOOLEAN DEFAULT 0")
            conn.exec_driver_sql("UPDATE employees SET manual_access = 1 WHERE position != 'Стажер'")
        if "manual_access_granted_date" not in employee_cols:
            conn.exec_driver_sql("ALTER TABLE employees ADD COLUMN manual_access_granted_date DATE")

        att_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(attestations)").fetchall()}
        if att_cols and "cooldown_until" not in att_cols:
            conn.exec_driver_sql("ALTER TABLE attestations ADD COLUMN cooldown_until DATE")


@app.on_event("startup")
def on_startup() -> None:
    Base.metadata.create_all(bind=engine)
    ensure_schema()
    with SessionLocal() as db:
        seed_data(db)


# -------------------------- auth pages / API --------------------------

LOGIN_ERROR_MESSAGES = {
    "wrong": "Неверный логин/email или пароль.",
    "blocked": "Аккаунт заблокирован. Обратитесь к главному администратору.",
    "locked": "Слишком много попыток входа. Подождите 10 минут и попробуйте снова.",
    "session": "Сессия истекла или недействительна. Войдите заново.",
    "required": "Для доступа к панели нужно войти.",
}


@app.get("/login")
def login_page(request: Request):
    token_user_id = parse_auth_token(request.cookies.get("hf_auth", ""))
    if token_user_id:
        with SessionLocal() as db:
            user = db.get(User, token_user_id)
            if user and user.status == "active":
                return redirect_to("/")
    code = request.query_params.get("error", "")
    return templates.TemplateResponse("login.html", {"request": request, "error": LOGIN_ERROR_MESSAGES.get(code, "")})


@app.post("/auth/login")
def auth_login(request: Request, login: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    identifier = login.strip()
    if login_locked_until(request, identifier):
        if wants_json(request):
            return JSONResponse({"detail": LOGIN_ERROR_MESSAGES["locked"]}, status_code=429)
        return login_error_redirect("locked")

    lowered = identifier.lower()
    user = db.query(User).filter(or_(func.lower(User.username) == lowered, func.lower(User.email) == lowered)).first()
    if not user or not verify_password(password, user.password_hash):
        register_login_failure(request, identifier)
        if wants_json(request):
            return JSONResponse({"detail": LOGIN_ERROR_MESSAGES["wrong"]}, status_code=401)
        return login_error_redirect("wrong")

    if user.status != "active":
        if wants_json(request):
            return JSONResponse({"detail": LOGIN_ERROR_MESSAGES["blocked"]}, status_code=403)
        return login_error_redirect("blocked")

    clear_login_failures(request, identifier)
    user.last_login_at = datetime.now()
    user.last_login = user.last_login_at
    user.updated_at = datetime.now()
    db.commit()

    if wants_json(request):
        response = JSONResponse({"ok": True, "user": user_public(user)})
    else:
        response = redirect_to("/")
    set_auth_cookie(response, user.id)
    return response


# Обратная совместимость со старой формой входа, если браузер закешировал /login как POST.
@app.post("/login")
def login_legacy(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    return auth_login(request, username, password, db)


@app.post("/auth/logout")
def auth_logout(request: Request):
    response = JSONResponse({"ok": True}) if wants_json(request) else redirect_to("/login")
    clear_auth_cookie(response)
    return response


@app.get("/logout")
def logout(request: Request):
    response = redirect_to("/login")
    clear_auth_cookie(response)
    return response


@app.get("/auth/me")
def auth_me(request: Request):
    user = require_user(request)
    return JSONResponse({"user": user_public(user)})


def validate_password_strength(password: str) -> Optional[str]:
    if len(password) < 10:
        return "Пароль должен быть минимум 10 символов."
    if not re.search(r"[A-ZА-Я]", password):
        return "Добавьте в пароль хотя бы одну заглавную букву."
    if not re.search(r"[a-zа-я]", password):
        return "Добавьте в пароль хотя бы одну строчную букву."
    if not re.search(r"\d", password):
        return "Добавьте в пароль хотя бы одну цифру."
    return None


@app.post("/auth/register")
def auth_register(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form("staff"),
    db: Session = Depends(get_db),
):
    actor = require_owner(request)
    username = username.strip()
    email = email.strip().lower()
    role = role if role in ROLE_LABELS else "staff"
    if actor.role != "owner" and role == "owner":
        raise HTTPException(403, "Только главный администратор может создавать owner")
    if db.query(User).filter(func.lower(User.username) == username.lower()).first():
        raise HTTPException(400, "Пользователь с таким логином уже существует")
    if db.query(User).filter(func.lower(User.email) == email.lower()).first():
        raise HTTPException(400, "Пользователь с таким email уже существует")
    strength_error = validate_password_strength(password)
    if strength_error:
        raise HTTPException(400, strength_error)
    user = User(username=username, email=email, password_hash=hash_password(password), role=role, status="active")
    db.add(user)
    add_history(db, None, "Пользователи", f"Создан пользователь {username} ({ROLE_LABELS.get(role, role)})", actor.username)
    db.commit()
    if wants_json(request):
        return JSONResponse({"ok": True, "user": user_public(user)})
    return redirect_to("/settings?users=created")


@app.patch("/auth/change-password")
@app.post("/auth/change-password")
def auth_change_password(
    request: Request,
    current_password: str = Form(""),
    new_password: str = Form(...),
    user_id: int = Form(0),
    db: Session = Depends(get_db),
):
    actor = require_user(request)
    target = actor
    if user_id:
        if actor.role not in OWNER_ROLES:
            raise HTTPException(403, "Только главный администратор может менять пароли других пользователей")
        found = db.get(User, user_id)
        if not found:
            raise HTTPException(404, "Пользователь не найден")
        target = found
    elif not verify_password(current_password, actor.password_hash):
        raise HTTPException(400, "Текущий пароль указан неверно")

    strength_error = validate_password_strength(new_password)
    if strength_error:
        raise HTTPException(400, strength_error)
    target.password_hash = hash_password(new_password)
    target.updated_at = datetime.now()
    add_history(db, None, "Пользователи", f"Изменен пароль пользователя {target.username}", actor.username)
    db.commit()
    if wants_json(request):
        return JSONResponse({"ok": True})
    return redirect_to("/settings?users=password_changed")


def block_user_common(user_id: int, request: Request, db: Session, block: bool):
    actor = require_owner(request)
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(404, "Пользователь не найден")
    if target.id == actor.id and block:
        raise HTTPException(400, "Нельзя заблокировать самого себя")
    if target.role == "owner" and actor.role != "owner":
        raise HTTPException(403, "Только owner может менять статус owner")
    target.status = "blocked" if block else "active"
    target.updated_at = datetime.now()
    add_history(db, None, "Пользователи", f"{'Заблокирован' if block else 'Разблокирован'} пользователь {target.username}", actor.username)
    db.commit()
    if wants_json(request):
        return JSONResponse({"ok": True, "user": user_public(target)})
    return redirect_to("/settings?users=updated")


@app.patch("/auth/block/{user_id}")
@app.post("/auth/block/{user_id}")
def auth_block_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    return block_user_common(user_id, request, db, True)


@app.patch("/auth/unblock/{user_id}")
@app.post("/auth/unblock/{user_id}")
def auth_unblock_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    return block_user_common(user_id, request, db, False)


@app.post("/auth/update-role/{user_id}")
def auth_update_role(user_id: int, request: Request, role: str = Form(...), db: Session = Depends(get_db)):
    actor = require_owner(request)
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(404, "Пользователь не найден")
    if target.id == actor.id and role != "owner":
        raise HTTPException(400, "Нельзя понизить самого себя")
    if role not in ROLE_LABELS:
        raise HTTPException(400, "Неизвестная роль")
    target.role = role
    # При смене роли сбрасываем ручные права, чтобы применились права новой роли.
    target.permissions = ""
    target.updated_at = datetime.now()
    add_history(db, None, "Пользователи", f"Изменена роль пользователя {target.username}: {ROLE_LABELS.get(role, role)}", actor.username)
    db.commit()
    return redirect_to("/settings?users=role_updated")


@app.post("/auth/update-permissions/{user_id}")
def auth_update_permissions(user_id: int, request: Request, permissions: list[str] = Form([]), db: Session = Depends(get_db)):
    actor = require_owner(request)
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(404, "Пользователь не найден")
    if target.role == "owner":
        raise HTTPException(400, "Права owner нельзя ограничить")
    save_user_permissions(target, permissions)
    target.updated_at = datetime.now()
    add_history(db, None, "Права доступа", f"Обновлены права пользователя {target.username}", actor.username)
    db.commit()
    return redirect_to("/settings?users=permissions_updated")


@app.post("/auth/reset-permissions/{user_id}")
def auth_reset_permissions(user_id: int, request: Request, db: Session = Depends(get_db)):
    actor = require_owner(request)
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(404, "Пользователь не найден")
    if target.role != "owner":
        target.permissions = ""
        target.updated_at = datetime.now()
        add_history(db, None, "Права доступа", f"Права пользователя {target.username} сброшены до роли {ROLE_LABELS.get(target.role, target.role)}", actor.username)
        db.commit()
    return redirect_to("/settings?users=permissions_reset")


# -------------------------- pages --------------------------

@app.get("/")
def dashboard(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_dashboard")
    maintenance_update(db)
    today_date = date.today()
    month = today_date.month
    year = today_date.year

    total_employees = db.query(Employee).filter(Employee.is_archived.is_(False)).count()
    on_vacation = db.query(Employee).filter(Employee.is_archived.is_(False), Employee.status == "В отпуске").count()
    active_warnings = db.query(Punishment).filter(Punishment.status == "Активно", Punishment.type == "Предупреждение").count()
    active_reprimands = db.query(Punishment).filter(Punishment.status == "Активно", Punishment.type == "Выговор").count()
    promoted_month = db.query(Promotion).filter(extract("month", Promotion.promotion_date) == month, extract("year", Promotion.promotion_date) == year).count()
    archived_month = db.query(Employee).filter(Employee.is_archived.is_(True), extract("month", Employee.removal_date) == month, extract("year", Employee.removal_date) == year).count()
    active_chskp = active_chskp_query(db).count()
    active_chsp = active_chsp_query(db).count()
    manual_reminders = [e for e in db.query(Employee).filter(Employee.is_archived.is_(False), Employee.position == "Стажер", Employee.manual_access.is_(False)).order_by(Employee.accepted_date.asc()).all() if manual_due(e)]
    notices = notifications_data(db)

    return render(
        request,
        "dashboard.html",
        {
            "page": "dashboard",
            "cards": {
                "Всего сотрудников": total_employees,
                "В отпуске": on_vacation,
                "Активных предупреждений": active_warnings,
                "Активных выговоров": active_reprimands,
                "Повышено за месяц": promoted_month,
                "Снято за месяц": archived_month,
                "В ЧСКП активно": active_chskp,
                "В ЧСП активно": active_chsp,
                "Мануал к выдаче": len(manual_reminders),
            },
            "soon_punishments": notices["punishments"],
            "soon_vacations": notices["vacations"],
            "manual_reminders": manual_reminders,
        },
    )


@app.get("/employees")
def employees_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_employees")
    maintenance_update(db)
    employees = db.query(Employee).filter(Employee.is_archived.is_(False)).all()
    employees = sorted(employees, key=position_sort_key)
    counts = employee_counts(db)
    matches = {e.id: find_blacklist_matches(db, e.nick, e.discord, e.discord_id, e.telegram, e.telegram_id, e.email) for e in employees}
    return render(request, "employees.html", {"page": "employees", "employees": employees, "counts": counts, "options": all_options(db), "chskp_matches": matches})


@app.get("/employees/{employee_id}/profile")
def employee_profile(employee_id: int, request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_employees")
    employee = (
        db.query(Employee)
        .options(joinedload(Employee.punishments), joinedload(Employee.vacations), joinedload(Employee.promotions), joinedload(Employee.histories))
        .filter(Employee.id == employee_id)
        .first()
    )
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    counts = employee_counts(db).get(employee.id, {"warnings": 0, "reprimands": 0})
    matches = find_blacklist_matches(db, employee.nick, employee.discord, employee.discord_id, employee.telegram, employee.telegram_id, employee.email)
    histories = db.query(ActionHistory).filter(ActionHistory.employee_id == employee.id).order_by(ActionHistory.created_at.desc()).all()
    return render(request, "employee_profile.html", {"page": "employees", "employee": employee, "counts": counts, "matches": matches, "histories": histories, "options": all_options(db)})


@app.get("/punishments")
def punishments_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_punishments")
    maintenance_update(db)
    employees = db.query(Employee).filter(Employee.is_archived.is_(False)).order_by(Employee.nick.asc()).all()
    punishments = db.query(Punishment).join(Employee).order_by(Punishment.created_at.desc()).all()
    return render(
        request,
        "punishments.html",
        {"page": "punishments", "employees": employees, "punishments": punishments, "options": all_options(db), "warning_days": get_int_setting(db, "warning_days", 7), "reprimand_days": get_int_setting(db, "reprimand_days", 14)},
    )


@app.get("/removals")
def removals_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_removals")
    maintenance_update(db)
    punishments = db.query(Punishment).join(Employee).order_by(Punishment.remove_date.asc()).all()
    return render(request, "removals.html", {"page": "removals", "punishments": punishments})


@app.get("/vacations")
def vacations_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_vacations")
    maintenance_update(db)
    employees = db.query(Employee).filter(Employee.is_archived.is_(False)).order_by(Employee.nick.asc()).all()
    vacations = db.query(Vacation).join(Employee).order_by(Vacation.created_at.desc()).all()
    return render(request, "vacations.html", {"page": "vacations", "employees": employees, "vacations": vacations, "options": all_options(db)})


@app.get("/promotions")
def promotions_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_promotions")
    employees = db.query(Employee).filter(Employee.is_archived.is_(False)).order_by(Employee.nick.asc()).all()
    promotions = db.query(Promotion).join(Employee).order_by(Promotion.created_at.desc()).all()
    return render(request, "promotions.html", {"page": "promotions", "employees": employees, "promotions": promotions, "options": all_options(db)})


@app.get("/attestation")
def attestation_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_attestation")
    maintenance_update(db)
    employees = db.query(Employee).filter(Employee.is_archived.is_(False)).order_by(Employee.nick.asc()).all()
    attestations = db.query(Attestation).join(Employee).order_by(Attestation.created_at.desc()).all()
    history = db.query(AttestationHistory).join(Employee).order_by(AttestationHistory.promoted_at.desc()).limit(800).all()
    return render(request, "attestation.html", {"page": "attestation", "employees": employees, "attestations": attestations, "history": history, "options": all_options(db), "manual_levels": MANUAL_LEVELS, "attestation_targets": ATTESTATION_TARGETS})


@app.get("/interviews")
def interviews_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_interviews")
    rows = db.query(Interview).order_by(Interview.created_at.desc()).all()
    interviewers = option_values(db, "interviewers")
    return render(request, "interviews.html", {"page": "interviews", "rows": rows, "interviewers": interviewers})


@app.get("/archive")
def archive_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_archive")
    employees = db.query(Employee).filter(Employee.is_archived.is_(True)).order_by(Employee.removal_date.desc()).all()
    return render(request, "archive.html", {"page": "archive", "employees": employees})


def render_blacklist_page(request: Request, db: Session, list_type: str, page: str):
    entries = db.query(BlacklistEntry).filter(BlacklistEntry.list_type == list_type).order_by(BlacklistEntry.created_at.desc()).all()
    employees = db.query(Employee).order_by(Employee.nick.asc()).all()
    title = f"{list_type} — {BLACKLIST_FULL[list_type]}"
    return render(request, "chskp.html", {"page": page, "entries": entries, "employees": employees, "options": all_options(db), "list_type": list_type, "title_text": title, "full_text": BLACKLIST_FULL[list_type], "other_page": "chsp" if page == "chskp" else "chskp", "other_type": "ЧСП" if page == "chskp" else "ЧСКП"})


@app.get("/chskp")
def chskp_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_chskp")
    return render_blacklist_page(request, db, "ЧСКП", "chskp")


@app.get("/chsp")
def chsp_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_chsp")
    return render_blacklist_page(request, db, "ЧСП", "chsp")


@app.get("/notifications")
def notifications_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_notifications")
    notices = notifications_data(db)
    return render(request, "notifications.html", {"page": "notifications", "soon_punishments": notices["punishments"], "soon_vacations": notices["vacations"]})


@app.get("/history")
def history_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_history")
    q = db.query(ActionHistory).outerjoin(Employee)
    date_from = request.query_params.get("date_from", "")
    date_to = request.query_params.get("date_to", "")
    action = request.query_params.get("action", "")
    text = request.query_params.get("text", "")
    if date_from:
        q = q.filter(ActionHistory.created_at >= datetime.combine(parse_date(date_from), time.min))
    if date_to:
        q = q.filter(ActionHistory.created_at <= datetime.combine(parse_date(date_to), time.max))
    if action:
        q = q.filter(ActionHistory.action_type.ilike(f"%{action}%"))
    if text:
        like = f"%{text}%"
        q = q.filter(or_(ActionHistory.details.ilike(like), ActionHistory.actor.ilike(like), Employee.nick.ilike(like)))
    rows = q.order_by(ActionHistory.created_at.desc()).limit(800).all()
    return render(request, "history.html", {"page": "history", "rows": rows, "filters": {"date_from": date_from, "date_to": date_to, "action": action, "text": text}})


@app.get("/settings")
def settings_page(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_settings")
    options_by_category = {
        "positions": db.query(SettingOption).filter(SettingOption.category == "positions").order_by(SettingOption.sort_order, SettingOption.value).all(),
        "accepted_by": db.query(SettingOption).filter(SettingOption.category == "accepted_by").order_by(SettingOption.sort_order, SettingOption.value).all(),
        "punishment_issuers": db.query(SettingOption).filter(SettingOption.category == "punishment_issuers").order_by(SettingOption.sort_order, SettingOption.value).all(),
        "interviewers": db.query(SettingOption).filter(SettingOption.category == "interviewers").order_by(SettingOption.sort_order, SettingOption.value).all(),
        "statuses": db.query(SettingOption).filter(SettingOption.category == "statuses").order_by(SettingOption.sort_order, SettingOption.value).all(),
    }
    settings = {"warning_days": get_int_setting(db, "warning_days", 7), "reprimand_days": get_int_setting(db, "reprimand_days", 14)}
    users = db.query(User).order_by(User.role.desc(), User.username.asc()).all()
    return render(request, "settings.html", {"page": "settings", "options_by_category": options_by_category, "settings": settings, "users": users, "roles": ROLE_LABELS, "permission_groups": PERMISSION_GROUPS, "user_permissions_list": editable_user_permissions})



# -------------------------- portable JSON backups --------------------------
BACKUP_MODELS = [
    User,
    SettingOption,
    AppSetting,
    Employee,
    Punishment,
    Vacation,
    Promotion,
    BlacklistEntry,
    Attestation,
    AttestationHistory,
    Interview,
    ActionHistory,
]
BACKUP_MODEL_BY_TABLE = {model.__tablename__: model for model in BACKUP_MODELS}
BACKUP_INSERT_ORDER = BACKUP_MODELS
BACKUP_DELETE_ORDER = list(reversed(BACKUP_MODELS))


def serialize_backup_value(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def parse_backup_value(column, value):
    if value in (None, ""):
        return None if column.nullable else value
    try:
        py_type = column.type.python_type
    except NotImplementedError:
        return value
    if py_type is datetime and isinstance(value, str):
        return datetime.fromisoformat(value)
    if py_type is date and isinstance(value, str):
        return date.fromisoformat(value)
    if py_type is time and isinstance(value, str):
        return time.fromisoformat(value)
    if py_type is bool:
        return parse_bool_value(str(value), default=bool(value))
    if py_type is int and value != "":
        return int(value)
    return value


def model_rows_for_backup(db: Session) -> dict:
    data = {}
    for model in BACKUP_MODELS:
        table = model.__tablename__
        rows = []
        for obj in db.query(model).all():
            row = {}
            for column in model.__table__.columns:
                row[column.name] = serialize_backup_value(getattr(obj, column.name))
            rows.append(row)
        data[table] = rows
    return data


def write_json_backup(db: Session, target: Path) -> None:
    payload = {
        "format": "holyfake-json-backup",
        "version": 2,
        "created_at": datetime.now().isoformat(),
        "tables": model_rows_for_backup(db),
    }
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def read_json_backup(path: Path) -> dict:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("format") != "holyfake-json-backup" or "tables" not in payload:
        raise ValueError("Это не резервная копия HolyFake JSON")
    return payload["tables"]


def read_sqlite_backup(path: Path) -> dict:
    result = {}
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    try:
        integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise ValueError("SQLite-файл поврежден")
        existing_tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        for table, model in BACKUP_MODEL_BY_TABLE.items():
            if table not in existing_tables:
                result[table] = []
                continue
            rows = []
            for row in conn.execute(f'SELECT * FROM "{table}"').fetchall():
                rows.append({k: row[k] for k in row.keys()})
            result[table] = rows
    finally:
        conn.close()
    return result


def restore_tables_from_backup(db: Session, tables: dict) -> None:
    # Полное восстановление: текущие данные заменяются данными из резервной копии.
    for model in BACKUP_DELETE_ORDER:
        db.query(model).delete(synchronize_session=False)
    db.flush()

    for model in BACKUP_INSERT_ORDER:
        table_rows = tables.get(model.__tablename__, [])
        if not table_rows:
            continue
        columns = {column.name: column for column in model.__table__.columns}
        prepared = []
        for source_row in table_rows:
            row = {}
            for name, column in columns.items():
                if name in source_row:
                    row[name] = parse_backup_value(column, source_row[name])
            prepared.append(row)
        if prepared:
            db.execute(model.__table__.insert(), prepared)
    db.flush()

    # Для PostgreSQL после ручной вставки id нужно обновить sequence, иначе новые записи могут конфликтовать.
    if not IS_SQLITE:
        for model in BACKUP_INSERT_ORDER:
            table = model.__tablename__
            if "id" not in model.__table__.columns:
                continue
            db.execute(text(f"SELECT setval(pg_get_serial_sequence('{table}', 'id'), COALESCE((SELECT MAX(id) FROM {table}), 1), (SELECT COUNT(*) FROM {table}) > 0)"))

@app.get("/backups")
def backups_page(request: Request):
    require_permission(request, "view_backups")
    files = sorted([*BACKUP_DIR.glob("*.json"), *BACKUP_DIR.glob("*.sqlite3")], key=lambda p: p.stat().st_mtime, reverse=True)
    backups = [
        {
            "name": f.name,
            "size_kb": round(f.stat().st_size / 1024, 1),
            "created": datetime.fromtimestamp(f.stat().st_mtime).strftime("%d.%m.%Y %H:%M"),
            "type": "JSON" if f.suffix.lower() == ".json" else "SQLite",
        }
        for f in files
    ]
    return render(request, "backups.html", {"page": "backups", "backups": backups, "params": request.query_params})


@app.get("/import-export")
def import_export_page(request: Request):
    require_permission(request, "view_import_export")
    return render(request, "import_export.html", {"page": "import_export", "params": request.query_params})

# -------------------------- employee actions --------------------------

@app.post("/employees/add")
def add_employee(
    request: Request,
    nick: str = Form(...),
    position: str = Form(...),
    discord: str = Form(""),
    discord_id: str = Form(""),
    telegram: str = Form(""),
    telegram_id: str = Form(""),
    email: str = Form(""),
    accepted_by: str = Form(""),
    accepted_date: str = Form(""),
    status: str = Form("Активен"),
    two_fa_enabled: str = Form("0"),
    manual_access: str = Form("0"),
    db: Session = Depends(get_db),
):
    user = require_permission(request, "manage_employees")
    nick = nick.strip()
    if db.query(Employee).filter(func.lower(Employee.nick) == nick.lower()).first():
        raise HTTPException(status_code=400, detail="Сотрудник с таким ником уже существует")
    matches = find_blacklist_matches(db, nick, discord, discord_id, telegram, telegram_id, email)
    employee = Employee(
        nick=nick,
        position=position,
        discord=discord.strip(),
        discord_id=discord_id.strip(),
        telegram=telegram.strip(),
        telegram_id=telegram_id.strip(),
        email=email.strip(),
        accepted_by=accepted_by,
        accepted_date=parse_date(accepted_date, date.today()),
        status=status,
        two_fa_enabled=parse_bool_value(two_fa_enabled),
        manual_access=parse_bool_value(manual_access, default=(position not in {"Стажер", "-", ""})),
        manual_access_granted_date=date.today() if parse_bool_value(manual_access, default=(position not in {"Стажер", "-", ""})) else None,
    )
    db.add(employee)
    db.flush()
    add_history(db, employee.id, "Добавление сотрудника", f"Добавлен сотрудник {employee.nick}", accepted_by or user.username)
    if matches:
        details = "; ".join([f"совпадение с {m['entry'].list_type} #{m['entry'].id}: {', '.join(m['fields'])}" for m in matches])
        add_history(db, employee.id, f"Предупреждение {blacklist_action_label(matches)}", details, "Система")
    db.commit()
    suffix = "?chskp_warning=1" if matches else ""
    return redirect_to(f"/employees{suffix}")


@app.post("/employees/{employee_id}/edit")
def edit_employee(
    employee_id: int,
    request: Request,
    nick: str = Form(...),
    position: str = Form(...),
    discord: str = Form(""),
    discord_id: str = Form(""),
    telegram: str = Form(""),
    telegram_id: str = Form(""),
    email: str = Form(""),
    accepted_by: str = Form(""),
    accepted_date: str = Form(""),
    status: str = Form("Активен"),
    two_fa_enabled: str = Form("0"),
    manual_access: str = Form("0"),
    db: Session = Depends(get_db),
):
    user = require_permission(request, "manage_employees")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    duplicate = db.query(Employee).filter(func.lower(Employee.nick) == nick.lower(), Employee.id != employee_id).first()
    if duplicate:
        raise HTTPException(status_code=400, detail="Сотрудник с таким ником уже существует")

    matches = find_blacklist_matches(db, nick, discord, discord_id, telegram, telegram_id, email)
    employee.nick = nick.strip()
    employee.position = position
    employee.discord = discord.strip()
    employee.discord_id = discord_id.strip()
    employee.telegram = telegram.strip()
    employee.telegram_id = telegram_id.strip()
    employee.email = email.strip()
    employee.accepted_by = accepted_by
    employee.accepted_date = parse_date(accepted_date, employee.accepted_date)
    employee.status = status
    employee.two_fa_enabled = parse_bool_value(two_fa_enabled)
    new_manual_access = parse_bool_value(manual_access, default=employee.manual_access)
    if new_manual_access and not employee.manual_access:
        employee.manual_access_granted_date = date.today()
    if not new_manual_access:
        employee.manual_access_granted_date = None
    employee.manual_access = new_manual_access
    add_history(db, employee.id, "Редактирование", "Данные сотрудника обновлены", accepted_by or user.username)
    if matches:
        details = "; ".join([f"совпадение с {m['entry'].list_type} #{m['entry'].id}: {', '.join(m['fields'])}" for m in matches])
        add_history(db, employee.id, f"Предупреждение {blacklist_action_label(matches)}", details, "Система")
    db.commit()
    suffix = "?chskp_warning=1" if matches else ""
    return redirect_to(f"/employees{suffix}")


@app.post("/employees/{employee_id}/archive")
def archive_employee(
    employee_id: int,
    request: Request,
    removal_reason: str = Form(...),
    removed_by: str = Form(...),
    removal_date: str = Form(""),
    removal_comment: str = Form(""),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_employees")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    employee.is_archived = True
    employee.status = "Снят"
    employee.removal_reason = removal_reason.strip()
    employee.removed_by = removed_by.strip()
    employee.removal_date = parse_date(removal_date, date.today())
    employee.removal_comment = removal_comment.strip()
    add_history(db, employee.id, "Снятие сотрудника", f"Причина: {employee.removal_reason}. Комментарий: {employee.removal_comment}", employee.removed_by)
    linked_user = db.query(User).filter(func.lower(User.username) == employee.nick.lower()).first()
    if linked_user and linked_user.status == "active" and linked_user.role != "owner":
        linked_user.status = "blocked"
        linked_user.updated_at = datetime.now()
        add_history(db, employee.id, "Автоблокировка доступа", f"Пользователь сайта {linked_user.username} заблокирован после снятия сотрудника", "Система")
    db.commit()
    return redirect_to("/employees")


@app.post("/employees/{employee_id}/delete")
def delete_employee_permanent(employee_id: int, request: Request, db: Session = Depends(get_db)):
    require_permission(request, "manage_employees")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    # Удаление намеренно НЕ записывается в журнал действий по требованию владельца панели.
    db.delete(employee)
    db.commit()
    return redirect_to("/employees")


@app.post("/archive/{employee_id}/restore")
def restore_employee(employee_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_permission(request, "manage_employees")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    employee.is_archived = False
    employee.status = "Активен"
    add_history(db, employee.id, "Возврат из архива", "Сотрудник восстановлен в активную таблицу", user.username)
    db.commit()
    return redirect_to("/archive")



@app.post("/employees/{employee_id}/manual/give")
def give_employee_manual(employee_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_permission(request, "manage_employees")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")

    employee.manual_access = True
    employee.two_fa_enabled = True
    employee.manual_access_granted_date = date.today()
    employee.updated_at = datetime.now()
    add_history(db, employee.id, "Выдача мануала", "Выдан доступ к мануалу и автоматически поставлена галочка 2FA", user.username)
    db.commit()

    return JSONResponse({"success": True, "manual_access": True, "two_fa_enabled": True})

@app.get("/api/employees/{employee_id}")
def employee_json(employee_id: int, request: Request, db: Session = Depends(get_db)):
    require_permission(request, "view_employees")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    return JSONResponse(
        {
            "id": employee.id,
            "nick": employee.nick,
            "position": employee.position,
            "discord": employee.discord,
            "discord_id": employee.discord_id,
            "telegram": employee.telegram,
            "telegram_id": employee.telegram_id,
            "email": employee.email,
            "accepted_by": employee.accepted_by,
            "accepted_date": employee.accepted_date.isoformat() if employee.accepted_date else "",
            "status": employee.status,
            "two_fa_enabled": "1" if employee.two_fa_enabled else "0",
            "manual_access": "1" if employee.manual_access else "0",
        }
    )


@app.get("/api/chskp/check")
def chskp_check(
    request: Request,
    nick: str = "",
    discord: str = "",
    discord_id: str = "",
    telegram: str = "",
    telegram_id: str = "",
    email: str = "",
    db: Session = Depends(get_db),
):
    require_permission(request, "view_employees")
    matches = find_blacklist_matches(db, nick, discord, discord_id, telegram, telegram_id, email)
    return JSONResponse(
        {
            "has_matches": bool(matches),
            "matches": [
                {
                    "id": m["entry"].id,
                    "list_type": m["entry"].list_type,
                    "nick": m["entry"].nick,
                    "fields": m["fields"],
                    "reason": m["entry"].reason,
                    "term": "Бессрочно" if m["entry"].term_type == "Бессрочно" else f"до {m['entry'].expires_at.strftime('%d.%m.%Y') if m['entry'].expires_at else '-'}",
                }
                for m in matches
            ],
        }
    )


# -------------------------- punishments --------------------------

@app.post("/punishments/add")
def add_punishment(
    request: Request,
    employee_id: int = Form(...),
    type: str = Form(...),
    reason: str = Form(...),
    issued_date: str = Form(""),
    issued_time: str = Form(""),
    issued_by: str = Form(...),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_punishments")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    start = parse_date(issued_date, date.today())
    days = get_int_setting(db, "warning_days", 7) if type == "Предупреждение" else get_int_setting(db, "reprimand_days", 14)
    punishment = Punishment(
        employee_id=employee.id,
        type=type,
        reason=reason.strip(),
        issued_date=start,
        issued_time=parse_time(issued_time),
        issued_by=issued_by,
        remove_date=start + timedelta(days=days),
        status="Активно",
    )
    db.add(punishment)
    add_history(db, employee.id, "Выдача наказания", f"{type}: {reason}. Снятие: {punishment.remove_date.strftime('%d.%m.%Y')}", issued_by)
    db.commit()
    return redirect_to("/punishments")


@app.post("/punishments/{punishment_id}/manual-remove")
def remove_punishment_manual(punishment_id: int, request: Request, actor: str = Form(""), db: Session = Depends(get_db)):
    user = require_permission(request, "manage_punishments")
    punishment = db.get(Punishment, punishment_id)
    if not punishment:
        raise HTTPException(404, "Наказание не найдено")
    punishment.status = "Снято вручную"
    punishment.removed_at = datetime.now()
    add_history(db, punishment.employee_id, "Ручное снятие наказания", f"{punishment.type}: {punishment.reason}", actor or user.username)
    db.commit()
    return redirect_to("/removals")


# -------------------------- vacations --------------------------

@app.post("/vacations/add")
def add_vacation(
    request: Request,
    employee_id: int = Form(...),
    reason: str = Form(...),
    start_date: str = Form(...),
    days: int = Form(...),
    issued_by: str = Form(...),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_vacations")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    start = parse_date(start_date, date.today())
    end = start + timedelta(days=max(days, 1))
    vacation = Vacation(employee_id=employee.id, reason=reason.strip(), start_date=start, days=max(days, 1), end_date=end, issued_by=issued_by, comment=comment.strip(), status="Активно")
    db.add(vacation)
    employee.status = "В отпуске"
    add_history(db, employee.id, "Выдача отпуска", f"{reason}. До: {end.strftime('%d.%m.%Y')}. Комментарий: {comment}", issued_by)
    db.commit()
    return redirect_to("/vacations")


@app.post("/vacations/{vacation_id}/finish")
def finish_vacation(vacation_id: int, request: Request, actor: str = Form(""), db: Session = Depends(get_db)):
    user = require_permission(request, "manage_vacations")
    vacation = db.get(Vacation, vacation_id)
    if not vacation:
        raise HTTPException(404, "Отпуск не найден")
    vacation.status = "Завершен вручную"
    employee = db.get(Employee, vacation.employee_id)
    if employee and not employee.is_archived:
        has_other_active = db.query(Vacation).filter(Vacation.employee_id == employee.id, Vacation.id != vacation.id, Vacation.status == "Активно").count()
        if not has_other_active:
            employee.status = "Активен"
    add_history(db, vacation.employee_id, "Ручное завершение отпуска", f"Отпуск завершен вручную: {vacation.reason}", actor or user.username)
    db.commit()
    return redirect_to("/vacations")


# -------------------------- promotions --------------------------

@app.post("/promotions/add")
def add_promotion(
    request: Request,
    employee_id: int = Form(...),
    new_position: str = Form(...),
    reason: str = Form(...),
    promoted_by: str = Form(...),
    promotion_date: str = Form(""),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_promotions")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    old = employee.position
    employee.position = new_position
    promotion = Promotion(employee_id=employee.id, old_position=old, new_position=new_position, reason=reason.strip(), promoted_by=promoted_by, promotion_date=parse_date(promotion_date, date.today()))
    db.add(promotion)
    add_history(db, employee.id, "Повышение", f"{old} → {new_position}. Причина: {reason}", promoted_by)
    db.commit()
    return redirect_to("/promotions")


# -------------------------- attestation / interviews --------------------------

@app.post("/attestation/add")
def add_attestation(
    request: Request,
    employee_id: int = Form(...),
    target_position: str = Form(...),
    manual_level: str = Form(...),
    created_by: str = Form(""),
    db: Session = Depends(get_db),
):
    user = require_permission(request, "manage_attestation")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    validate_attestation_target(employee, target_position)
    active_exists = db.query(Attestation).filter(Attestation.employee_id == employee.id, Attestation.status.in_(["Активно", "КД", "Задержка"])).first()
    if active_exists:
        raise HTTPException(400, "У сотрудника уже есть активная аттестация, КД или задержка")
    item = Attestation(employee_id=employee.id, target_position=target_position, manual_level=manual_level, remaining_attempts=3, status="Активно", created_by=created_by or user.username)
    db.add(item)
    add_history(db, employee.id, "Аттестация", f"Начата аттестация: {employee.position} → {target_position}. Мануал: {manual_level}", created_by or user.username)
    db.commit()
    return redirect_to("/attestation")


@app.post("/attestation/{attestation_id}/fail")
def fail_attestation(attestation_id: int, request: Request, actor: str = Form(""), db: Session = Depends(get_db)):
    user = require_permission(request, "manage_attestation")
    item = db.get(Attestation, attestation_id)
    if not item:
        raise HTTPException(404, "Аттестация не найдена")
    if item.status != "Активно":
        return redirect_to("/attestation")
    item.remaining_attempts = max((item.remaining_attempts or 0) - 1, 0)
    actor_name = actor or user.username
    if item.remaining_attempts <= 0:
        item.status = "Задержка"
        item.delay_until = date.today() + timedelta(days=7)
        item.cooldown_until = None
        add_history(db, item.employee_id, "Аттестация не сдана", f"3 провала подряд. Повторная аттестация доступна с {item.delay_until.strftime('%d.%m.%Y')}", actor_name)
    else:
        item.status = "КД"
        item.cooldown_until = date.today() + timedelta(days=2)
        add_history(db, item.employee_id, "Аттестация не сдана", f"Осталось попыток: {item.remaining_attempts}. Повторная попытка доступна с {item.cooldown_until.strftime('%d.%m.%Y')}", actor_name)
    db.commit()
    return redirect_to("/attestation")


@app.post("/attestation/{attestation_id}/pass")
def pass_attestation(attestation_id: int, request: Request, conducted_by: str = Form(...), comment: str = Form(""), db: Session = Depends(get_db)):
    require_permission(request, "manage_attestation")
    item = db.get(Attestation, attestation_id)
    if not item:
        raise HTTPException(404, "Аттестация не найдена")
    employee = db.get(Employee, item.employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    validate_attestation_target(employee, item.target_position)
    old_position = employee.position
    employee.position = item.target_position
    item.status = "Пройдено"
    item.remaining_attempts = 0
    item.cooldown_until = None
    item.delay_until = None
    history = AttestationHistory(employee_id=employee.id, promoted_to=item.target_position, promoted_at=datetime.now(), conducted_by=conducted_by, comment=comment.strip())
    db.add(history)
    db.add(Promotion(employee_id=employee.id, old_position=old_position, new_position=item.target_position, reason="Успешно пройдена аттестация", promoted_by=conducted_by, promotion_date=date.today()))
    add_history(db, employee.id, "Аттестация пройдена", f"{old_position} → {item.target_position}. Комментарий: {comment}", conducted_by)
    db.commit()
    return redirect_to("/attestation")


@app.post("/interviews/add")
def add_interview(
    request: Request,
    candidate_name: str = Form(...),
    server_nick: str = Form(...),
    age: int = Form(0),
    intelligence_score: int = Form(0),
    punishment_mode: str = Form("Нет"),
    punishment_history: str = Form(""),
    communication_feelings: str = Form(""),
    accepted: str = Form("Нет"),
    cool_score: int = Form(0),
    cheats_mode: str = Form("Нет"),
    cheats_info: str = Form(""),
    accepted_by: str = Form(""),
    db: Session = Depends(get_db),
):
    user = require_permission(request, "manage_interviews")
    punishment_text = "Нет" if punishment_mode == "Нет" else (punishment_history.strip() or "Есть, но описание не указано")
    cheats_text = "Нет" if cheats_mode == "Нет" else (cheats_info.strip() or "Есть, но название не указано")
    row = Interview(
        candidate_name=candidate_name.strip(),
        server_nick=server_nick.strip(),
        age=max(age, 0),
        intelligence_score=min(max(intelligence_score, 0), 10),
        punishment_history=punishment_text,
        communication_feelings=communication_feelings.strip(),
        accepted=accepted,
        cool_score=min(max(cool_score, 0), 5),
        cheats_info=cheats_text,
        accepted_by=accepted_by,
    )
    db.add(row)
    add_history(db, None, "Собеседование", f"Кандидат {row.server_nick}: принят — {accepted}", accepted_by or user.username)
    db.commit()
    return redirect_to("/interviews")


@app.post("/interviews/{interview_id}/delete")
def delete_interview(interview_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_permission(request, "manage_interviews")
    row = db.get(Interview, interview_id)
    if row:
        add_history(db, None, "Удаление собеседования", f"Удалена анкета {row.server_nick}", user.username)
        db.delete(row)
        db.commit()
    return redirect_to("/interviews")


# -------------------------- chskp --------------------------

@app.post("/employees/{employee_id}/chskp")
def add_employee_to_chskp(
    employee_id: int,
    request: Request,
    reason: str = Form(...),
    added_by: str = Form(...),
    added_date: str = Form(""),
    term_type: str = Form("Бессрочно"),
    expires_at: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_chskp")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    entry = BlacklistEntry(
        list_type="ЧСКП",
        nick=employee.nick,
        position=employee.position,
        discord=employee.discord,
        discord_id=employee.discord_id,
        telegram=employee.telegram,
        telegram_id=employee.telegram_id,
        email=employee.email,
        reason=reason.strip(),
        added_by=added_by.strip(),
        added_date=parse_date(added_date, date.today()),
        term_type=term_type,
        expires_at=parse_date(expires_at) if term_type == "До даты" else None,
        comment=comment.strip(),
    )
    db.add(entry)
    add_history(db, employee.id, "Добавление в ЧСКП", f"Сотрудник внесен в ЧСКП. Причина: {entry.reason}", entry.added_by)
    db.commit()
    return redirect_to(f"/employees/{employee_id}/profile")




@app.post("/employees/{employee_id}/chsp")
def add_employee_to_chsp(
    employee_id: int,
    request: Request,
    reason: str = Form(...),
    added_by: str = Form(...),
    added_date: str = Form(""),
    term_type: str = Form("Бессрочно"),
    expires_at: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_chsp")
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Сотрудник не найден")
    entry = BlacklistEntry(
        list_type="ЧСП",
        nick=employee.nick,
        position=employee.position,
        discord=employee.discord,
        discord_id=employee.discord_id,
        telegram=employee.telegram,
        telegram_id=employee.telegram_id,
        email=employee.email,
        reason=reason.strip(),
        added_by=added_by.strip(),
        added_date=parse_date(added_date, date.today()),
        term_type=term_type,
        expires_at=parse_date(expires_at) if term_type == "До даты" else None,
        comment=comment.strip(),
    )
    db.add(entry)
    add_history(db, employee.id, "Добавление в ЧСП", f"Сотрудник внесен в ЧСП. Причина: {entry.reason}", entry.added_by)
    db.commit()
    return redirect_to(f"/employees/{employee_id}/profile")


@app.post("/chskp/add")
def add_chskp(
    request: Request,
    nick: str = Form(""),
    position: str = Form(""),
    discord: str = Form(""),
    discord_id: str = Form(""),
    telegram: str = Form(""),
    telegram_id: str = Form(""),
    email: str = Form(""),
    reason: str = Form(...),
    added_by: str = Form(...),
    added_date: str = Form(""),
    term_type: str = Form("Бессрочно"),
    expires_at: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_chskp")
    entry = BlacklistEntry(
        list_type="ЧСКП",
        nick=nick.strip(),
        position=position.strip(),
        discord=discord.strip(),
        discord_id=discord_id.strip(),
        telegram=telegram.strip(),
        telegram_id=telegram_id.strip(),
        email=email.strip(),
        reason=reason.strip(),
        added_by=added_by.strip(),
        added_date=parse_date(added_date, date.today()),
        term_type=term_type,
        expires_at=parse_date(expires_at) if term_type == "До даты" else None,
        comment=comment.strip(),
    )
    db.add(entry)
    add_history(db, None, "Добавление в ЧСКП", f"{entry.nick or entry.discord or entry.discord_id}: {entry.reason}", entry.added_by)
    db.commit()
    return redirect_to("/chskp")




@app.post("/chsp/add")
def add_chsp(
    request: Request,
    nick: str = Form(""),
    position: str = Form(""),
    discord: str = Form(""),
    discord_id: str = Form(""),
    telegram: str = Form(""),
    telegram_id: str = Form(""),
    email: str = Form(""),
    reason: str = Form(...),
    added_by: str = Form(...),
    added_date: str = Form(""),
    term_type: str = Form("Бессрочно"),
    expires_at: str = Form(""),
    comment: str = Form(""),
    db: Session = Depends(get_db),
):
    require_permission(request, "manage_chsp")
    entry = BlacklistEntry(
        list_type="ЧСП",
        nick=nick.strip(),
        position=position.strip(),
        discord=discord.strip(),
        discord_id=discord_id.strip(),
        telegram=telegram.strip(),
        telegram_id=telegram_id.strip(),
        email=email.strip(),
        reason=reason.strip(),
        added_by=added_by.strip(),
        added_date=parse_date(added_date, date.today()),
        term_type=term_type,
        expires_at=parse_date(expires_at) if term_type == "До даты" else None,
        comment=comment.strip(),
    )
    db.add(entry)
    add_history(db, None, "Добавление в ЧСП", f"{entry.nick or entry.discord or entry.discord_id}: {entry.reason}", entry.added_by)
    db.commit()
    return redirect_to("/chsp")


@app.post("/chskp/{entry_id}/delete")
def delete_chskp(entry_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_permission(request, "manage_chskp")
    entry = db.get(BlacklistEntry, entry_id)
    if entry:
        add_history(db, None, "Удаление из ЧСКП", f"Удалена запись: {entry.nick or entry.discord or entry.discord_id}", user.username)
        db.delete(entry)
        db.commit()
    return redirect_to("/chskp")




@app.post("/chsp/{entry_id}/delete")
def delete_chsp(entry_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_permission(request, "manage_chsp")
    entry = db.get(BlacklistEntry, entry_id)
    if entry:
        add_history(db, None, "Удаление из ЧСП", f"Удалена запись: {entry.nick or entry.discord or entry.discord_id}", user.username)
        db.delete(entry)
        db.commit()
    return redirect_to("/chsp")


# -------------------------- settings --------------------------

@app.post("/settings/options/add")
def add_option(request: Request, category: str = Form(...), value: str = Form(...), db: Session = Depends(get_db)):
    require_permission(request, "manage_settings")
    value = value.strip()
    if value:
        exists = db.query(SettingOption).filter(SettingOption.category == category, func.lower(SettingOption.value) == value.lower()).first()
        if not exists:
            max_order = db.query(func.max(SettingOption.sort_order)).filter(SettingOption.category == category).scalar() or 0
            db.add(SettingOption(category=category, value=value, sort_order=max_order + 1))
            add_history(db, None, "Настройки", f"Добавлено значение {value} в {category}", request.state.current_user.username)
            db.commit()
    return redirect_to("/settings")


@app.post("/settings/options/{option_id}/delete")
def delete_option(option_id: int, request: Request, db: Session = Depends(get_db)):
    require_permission(request, "manage_settings")
    option = db.get(SettingOption, option_id)
    if option:
        add_history(db, None, "Настройки", f"Удалено значение {option.value} из {option.category}", request.state.current_user.username)
        db.delete(option)
        db.commit()
    return redirect_to("/settings")


@app.post("/settings/penalty-days")
def update_penalty_days(request: Request, warning_days: int = Form(...), reprimand_days: int = Form(...), db: Session = Depends(get_db)):
    require_permission(request, "manage_settings")
    set_setting(db, "warning_days", str(max(warning_days, 1)))
    set_setting(db, "reprimand_days", str(max(reprimand_days, 1)))
    add_history(db, None, "Настройки", f"Сроки наказаний обновлены: предупреждение {warning_days}, выговор {reprimand_days}", request.state.current_user.username)
    db.commit()
    return redirect_to("/settings")

# -------------------------- export / import / backups --------------------------

def append_sheet(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 45)


@app.get("/export/excel")
def export_excel(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "manage_import_export")
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    employees = db.query(Employee).order_by(Employee.is_archived.asc(), Employee.nick.asc()).all()
    counts = employee_counts(db)
    append_sheet(
        ws,
        ["Ник", "Должность", "Дискорд", "Discord ID", "Telegram", "Telegram ID", "Почта", "Предупреждения", "Выговоры", "2FA", "Доступ к мануалу", "Кто принял", "Дата принятия", "Статус", "Архив"],
        [
            [
                e.nick,
                e.position,
                e.discord,
                e.discord_id,
                e.telegram,
                e.telegram_id,
                e.email,
                counts.get(e.id, {"warnings": 0})["warnings"],
                counts.get(e.id, {"reprimands": 0})["reprimands"],
                "Да" if e.two_fa_enabled else "Нет",
                "Да" if e.manual_access else "Нет",
                e.accepted_by,
                e.accepted_date.strftime("%d.%m.%Y") if e.accepted_date else "",
                e.status,
                "Да" if e.is_archived else "Нет",
            ]
            for e in employees
        ],
    )

    sheets = [
        ("Наказания", ["Ник", "Тип", "Причина", "Дата выдачи", "Время", "Кто выдал", "Дата снятия", "Статус"], [
            [p.employee.nick, p.type, p.reason, p.issued_date.strftime("%d.%m.%Y"), p.issued_time.strftime("%H:%M") if p.issued_time else "", p.issued_by, p.remove_date.strftime("%d.%m.%Y"), p.status]
            for p in db.query(Punishment).join(Employee).order_by(Punishment.created_at.desc()).all()
        ]),
        ("Отпуска", ["Ник", "Причина", "Дата начала", "Дата окончания", "Дней", "Кто выдал", "Статус", "Комментарий"], [
            [v.employee.nick, v.reason, v.start_date.strftime("%d.%m.%Y"), v.end_date.strftime("%d.%m.%Y"), v.days, v.issued_by, v.status, v.comment]
            for v in db.query(Vacation).join(Employee).order_by(Vacation.created_at.desc()).all()
        ]),
        ("Повышения", ["Ник", "Старая должность", "Новая должность", "Причина", "Кто повысил", "Дата"], [
            [p.employee.nick, p.old_position, p.new_position, p.reason, p.promoted_by, p.promotion_date.strftime("%d.%m.%Y")]
            for p in db.query(Promotion).join(Employee).order_by(Promotion.created_at.desc()).all()
        ]),
        ("Черные списки", ["Тип", "Ник", "Должность", "Discord", "Discord ID", "Telegram", "Telegram ID", "Почта", "Причина", "Кто внес", "Дата", "Срок", "До", "Статус"], [
            [b.list_type, b.nick, b.position, b.discord, b.discord_id, b.telegram, b.telegram_id, b.email, b.reason, b.added_by, b.added_date.strftime("%d.%m.%Y"), b.term_type, b.expires_at.strftime("%d.%m.%Y") if b.expires_at else "", blacklist_status(b)]
            for b in db.query(BlacklistEntry).order_by(BlacklistEntry.list_type.asc(), BlacklistEntry.created_at.desc()).all()
        ]),
        ("Аттестация", ["Ник", "Текущая должность", "На должность", "Попыток", "Мануал", "Статус", "Задержка до"], [
            [a.employee.nick, a.employee.position, a.target_position, a.remaining_attempts, a.manual_level, a.status, a.delay_until.strftime("%d.%m.%Y") if a.delay_until else ""]
            for a in db.query(Attestation).join(Employee).order_by(Attestation.created_at.desc()).all()
        ]),
        ("История аттестаций", ["Ник", "Повышен до", "Дата и время", "Кто провел", "Комментарий"], [
            [h.employee.nick, h.promoted_to, h.promoted_at.strftime("%d.%m.%Y %H:%M"), h.conducted_by, h.comment]
            for h in db.query(AttestationHistory).join(Employee).order_by(AttestationHistory.promoted_at.desc()).all()
        ]),
        ("Собеседования", ["Дата", "Имя", "Ник", "Возраст", "Ум", "Наказания", "Ощущения", "Принят", "Крутость", "Читы", "Кто принял"], [
            [i.created_at.strftime("%d.%m.%Y %H:%M"), i.candidate_name, i.server_nick, i.age, i.intelligence_score, i.punishment_history, i.communication_feelings, i.accepted, i.cool_score, i.cheats_info, i.accepted_by]
            for i in db.query(Interview).order_by(Interview.created_at.desc()).all()
        ]),
        ("Журнал", ["Дата", "Сотрудник", "Действие", "Детали", "Кто"], [
            [h.created_at.strftime("%d.%m.%Y %H:%M"), h.employee.nick if h.employee else "-", h.action_type, h.details, h.actor]
            for h in db.query(ActionHistory).outerjoin(Employee).order_by(ActionHistory.created_at.desc()).limit(3000).all()
        ]),
    ]
    for title, headers, rows in sheets:
        append_sheet(wb.create_sheet(title), headers, rows)

    filename = f"holyfake_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = EXPORT_DIR / filename
    wb.save(path)
    add_history(db, None, "Экспорт Excel", f"Создан файл {filename}", request.state.current_user.username)
    db.commit()
    return FileResponse(path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def google_sheet_to_csv_url(url: str) -> str:
    url = url.strip()
    if "output=csv" in url or "format=csv" in url:
        return url
    parsed = urllib.parse.urlparse(url)
    gid = urllib.parse.parse_qs(parsed.query).get("gid", ["0"])[0]
    parts = parsed.path.split("/")
    if "d" in parts:
        try:
            sheet_id = parts[parts.index("d") + 1]
            return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
        except IndexError:
            pass
    return url


def normalize_key(value: str) -> str:
    value = (value or "").strip().lower().replace("ё", "е")
    value = value.replace("discord", "дискорд").replace("telegram", "тг")
    return re.sub(r"[^a-zа-я0-9]+", "", value)


def row_value(row: dict, *keys: str) -> str:
    lookup = {normalize_key(str(k)): v for k, v in row.items() if k is not None}
    for key in keys:
        compact = normalize_key(key)
        if compact in lookup:
            return str(lookup[compact] or "").strip()
    return ""


def parse_count_cell(value: str) -> int:
    text = str(value or "").strip()
    if not text:
        return 0
    m = re.search(r"\d+", text)
    return int(m.group(0)) if m else 0


def parse_bool_from_sheet(value: str, default: bool = False) -> bool:
    text = str(value or "").strip()
    # В Google Sheets в ячейке часто остается просто символ галочки/крестика или текст вроде TRUE/FALSE.
    return parse_bool_value(text, default=default)


def find_header_line(lines: list[str]) -> int:
    for i, line in enumerate(lines):
        low = normalize_key(line)
        if "ник" in low and ("должность" in low or "дискорд" in low):
            return i
    return 0


def dict_rows_from_csv(content: str) -> list[dict]:
    lines = content.splitlines()
    if not lines:
        return []
    start = find_header_line(lines)
    reader = csv.DictReader(io.StringIO("\n".join(lines[start:])))
    return list(reader)


def dict_rows_from_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_index = 0
    for i, row in enumerate(rows):
        joined = " ".join(str(c or "") for c in row)
        low = normalize_key(joined)
        if "ник" in low and ("должность" in low or "дискорд" in low):
            header_index = i
            break
    headers = [str(c or "").strip() for c in rows[header_index]]
    result = []
    for row in rows[header_index + 1:]:
        if not any(c not in (None, "") for c in row):
            continue
        item = {headers[i] if i < len(headers) and headers[i] else f"col_{i}": row[i] if i < len(row) else "" for i in range(max(len(headers), len(row)))}
        result.append(item)
    return result


def import_employee_rows(db: Session, rows: list[dict], actor: str) -> dict[str, int]:
    imported = skipped = warnings = created_punishments = 0
    for row in rows:
        nick = row_value(row, "ник", "nick", "nickname")
        if not nick:
            skipped += 1
            continue
        if db.query(Employee).filter(func.lower(Employee.nick) == nick.lower()).first():
            skipped += 1
            continue
        position = row_value(row, "должность", "position") or "Стажер"
        discord = row_value(row, "дискорд", "discord")
        discord_id = row_value(row, "discord id", "дискорд id", "дискорр id", "discordid", "дискордid")
        telegram = row_value(row, "telegram", "тг", "tg")
        telegram_id = row_value(row, "telegram id", "тг id", "tg id", "тгid")
        email = row_value(row, "почта", "email", "mail")
        accepted_by = row_value(row, "кто принял", "кто принял?", "accepted_by") or actor
        accepted_date = parse_date_flexible(row_value(row, "дата принятия", "дата принятия ", "accepted_date"), date.today())
        status = row_value(row, "статус", "status") or "Активен"
        two_fa_raw = row_value(row, "2fa", "2 fa", "2ФА", "привязка аккаунта", "привязка акка", "привязка")
        manual_raw = row_value(row, "доступ к мануалу", "доступ к мануалу на млку", "мануал", "manual access")
        two_fa = parse_bool_from_sheet(two_fa_raw, False)
        manual_default = position not in {"Стажер", "-", ""}
        manual_access_value = parse_bool_from_sheet(manual_raw, manual_default)
        matches = find_blacklist_matches(db, nick, discord, discord_id, telegram, telegram_id, email)
        if matches:
            warnings += 1
        employee = Employee(
            nick=nick.strip(),
            position=position.strip(),
            discord=discord.strip(),
            discord_id=discord_id.strip(),
            telegram=telegram.strip(),
            telegram_id=telegram_id.strip(),
            email=email.strip(),
            accepted_by=accepted_by.strip(),
            accepted_date=accepted_date,
            status=status.strip(),
            two_fa_enabled=two_fa,
            manual_access=manual_access_value,
            manual_access_granted_date=date.today() if manual_access_value else None,
        )
        db.add(employee)
        db.flush()

        warning_count = parse_count_cell(row_value(row, "предупреждения", "пред", "warnings"))
        reprimand_count = parse_count_cell(row_value(row, "выговоры", "выг", "reprimands"))
        for _ in range(warning_count):
            db.add(Punishment(employee_id=employee.id, type="Предупреждение", reason="Импортировано из таблицы", issued_date=date.today(), issued_time=datetime.now().time().replace(microsecond=0, second=0), issued_by=actor, remove_date=date.today() + timedelta(days=get_int_setting(db, "warning_days", 7)), status="Активно"))
            created_punishments += 1
        for _ in range(reprimand_count):
            db.add(Punishment(employee_id=employee.id, type="Выговор", reason="Импортировано из таблицы", issued_date=date.today(), issued_time=datetime.now().time().replace(microsecond=0, second=0), issued_by=actor, remove_date=date.today() + timedelta(days=get_int_setting(db, "reprimand_days", 14)), status="Активно"))
            created_punishments += 1

        add_history(db, employee.id, "Импорт таблицы", "Сотрудник импортирован из таблицы", actor)
        if matches:
            details = "; ".join([f"совпадение с {m['entry'].list_type} #{m['entry'].id}: {', '.join(m['fields'])}" for m in matches])
            add_history(db, employee.id, f"Предупреждение {blacklist_action_label(matches)}", details, "Система")
        imported += 1
    db.commit()
    return {"imported": imported, "skipped": skipped, "warnings": warnings, "punishments": created_punishments}


def import_result_redirect(stats: dict[str, int], error: str = "") -> RedirectResponse:
    params = urllib.parse.urlencode({k: v for k, v in stats.items()} | ({"error": error} if error else {}))
    return redirect_to(f"/import-export?{params}")


@app.post("/import/google-sheets")
def import_google_sheets(request: Request, google_url: str = Form(...), db: Session = Depends(get_db)):
    require_permission(request, "manage_import_export")
    csv_url = google_sheet_to_csv_url(google_url)
    try:
        with urllib.request.urlopen(csv_url, timeout=20) as response:
            content = response.read().decode("utf-8-sig")
    except Exception as exc:
        return import_result_redirect({"imported": 0, "skipped": 0, "warnings": 0, "punishments": 0}, f"Не удалось загрузить Google Sheets. Чаще всего это 401: таблица закрыта. Откройте доступ по ссылке или скачайте таблицу как .xlsx и импортируйте файлом. Детали: {exc}")
    rows = dict_rows_from_csv(content)
    stats = import_employee_rows(db, rows, request.state.current_user.username)
    return import_result_redirect(stats)


@app.post("/import/file")
async def import_table_file(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    require_permission(request, "manage_import_export")
    suffix = Path(file.filename or "").suffix.lower()
    content = await file.read()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                tmp_path = Path(tmp.name)
            try:
                rows = dict_rows_from_xlsx(tmp_path)
            finally:
                tmp_path.unlink(missing_ok=True)
        elif suffix in {".csv", ".txt"}:
            rows = dict_rows_from_csv(content.decode("utf-8-sig"))
        else:
            return import_result_redirect({"imported": 0, "skipped": 0, "warnings": 0, "punishments": 0}, "Поддерживаются только .xlsx, .xlsm и .csv")
    except Exception as exc:
        return import_result_redirect({"imported": 0, "skipped": 0, "warnings": 0, "punishments": 0}, f"Не удалось прочитать файл: {exc}")
    stats = import_employee_rows(db, rows, request.state.current_user.username)
    return import_result_redirect(stats)


@app.post("/backups/create")
def create_backup(request: Request, db: Session = Depends(get_db)):
    require_permission(request, "manage_backups")
    filename = f"holyfake_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    target = BACKUP_DIR / filename
    db.flush()
    write_json_backup(db, target)
    add_history(db, None, "Резервная копия", f"Создана JSON-копия {filename}", request.state.current_user.username)
    db.commit()
    return redirect_to("/backups")


@app.get("/backups/download/{filename}")
def download_backup(filename: str, request: Request):
    require_permission(request, "view_backups")
    safe_name = Path(filename).name
    path = BACKUP_DIR / safe_name
    if not path.exists() or path.suffix.lower() not in {".json", ".sqlite3"}:
        raise HTTPException(404, "Файл не найден")
    return FileResponse(path, filename=safe_name, media_type="application/octet-stream")


@app.post("/backups/restore")
async def restore_backup(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = require_permission(request, "manage_backups")
    content = await file.read()
    if not content:
        return redirect_to("/backups?restore_error=empty")

    suffix = Path(file.filename or "backup.json").suffix.lower()
    tmp_suffix = suffix if suffix in {".json", ".sqlite3"} else ".backup"
    with tempfile.NamedTemporaryFile(delete=False, suffix=tmp_suffix) as tmp:
        tmp.write(content)
        tmp_path = Path(tmp.name)

    try:
        # Перед восстановлением всегда создаем аварийную JSON-копию текущей базы.
        emergency_name = f"before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        write_json_backup(db, BACKUP_DIR / emergency_name)

        if suffix == ".sqlite3":
            tables = read_sqlite_backup(tmp_path)
        else:
            tables = read_json_backup(tmp_path)

        restore_tables_from_backup(db, tables)
        # Если в загруженной копии не было owner, сайт не должен остаться без главного доступа.
        owner = db.query(User).filter(func.lower(User.username) == "owner").first()
        if not owner:
            db.add(User(username="owner", email="owner@holyfake.local", password_hash=hash_password("HolyFake#2026!"), role="owner", status="active"))
        db.commit()
        return redirect_to("/backups?restored=1")
    except Exception as exc:
        db.rollback()
        return redirect_to(f"/backups?restore_error={urllib.parse.quote(str(exc)[:160])}")
    finally:
        tmp_path.unlink(missing_ok=True)


@app.get("/health")
def health():
    return {"status": "ok"}
